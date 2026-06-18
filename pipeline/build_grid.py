#!/usr/bin/env python3
"""Build grid.json from the Form S-6 'Land-Water ID' worksheet.

Reads the 840 grid vertices (21x40, ~3 statute-mile spacing) over southern
Florida and writes them as JSON for the Leaflet viewer.

Columns in the worksheet: E-W (mi), N-S (mi), Latitude, Longitude, ID
  ID = 0 -> water, ID = 1 -> land

Output: outputs/web/grid.json
"""
import json
import os
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
XLSX = os.path.join(ROOT, "FormS6Input.xlsx")
OUT = os.path.join(ROOT, "outputs", "web", "grid.json")

# Landfall / track reference (ROA p. 185)
LANDFALL = {"lat": 25.8611, "lon": -80.1196}


def main():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Land-Water ID"]

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    points = []
    ews, nss = set(), set()
    for r in rows:
        ew, ns, lat, lon, idv = r[0], r[1], r[2], r[3], r[4]
        if ew is None or lat is None:
            continue
        ew = int(round(ew)); ns = int(round(ns))
        land = int(idv) == 1
        points.append({
            "ew": ew, "ns": ns,
            "lat": round(float(lat), 6), "lon": round(float(lon), 6),
            "land": land,
        })
        ews.add(ew); nss.add(ns)

    n_land = sum(1 for p in points if p["land"])
    grid = {
        "n_points": len(points),
        "n_land": n_land,
        "n_water": len(points) - n_land,
        "ew_values": sorted(ews),
        "ns_values": sorted(nss),
        "landfall": LANDFALL,
        # storm track: due west from (0,0) to (117,0) over 12 hours
        "track": {"ew_start": 0, "ew_end": max(ews), "ns": 0, "hours": 12},
        "points": points,
    }

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w") as f:
        json.dump(grid, f)

    print(f"Wrote {OUT}")
    print(f"  points : {grid['n_points']} (expected 840)")
    print(f"  land   : {grid['n_land']} (expected 682)")
    print(f"  water  : {grid['n_water']}")
    print(f"  E-W    : {min(ews)}..{max(ews)} ({len(ews)} cols, expected 40)")
    print(f"  N-S    : {min(nss)}..{max(nss)} ({len(nss)} rows, expected 21)")


if __name__ == "__main__":
    main()
