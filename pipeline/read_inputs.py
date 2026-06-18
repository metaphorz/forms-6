#!/usr/bin/env python3
"""Extract the Form S-6 input vectors from FormS6Input.xlsx → inputs.json.

Uses the 'SA all Variables' worksheet, which holds all 100 input vectors with
the full variable set (CP, Rmax, VT, WSP, CF, FFP, Quantile) for each of the
three hurricane categories (1, 3, 5).

Layout (per the worksheet header rows):
  col A  = Vector number
  cols B..H  = Cat 1: CP, Rmax, VT, WSP, CF, FFP, Quantile
  col  I = blank
  cols J..P  = Cat 3: CP, Rmax, VT, WSP, CF, FFP, Quantile
  col  Q = blank
  cols R..X  = Cat 5: CP, Rmax, VT, WSP, CF, FFP, Quantile

Output: outputs/web/inputs.json
  { "vars": [...], "cat1": [{...100}], "cat3": [...], "cat5": [...] }
"""
import json
import os
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
XLSX = os.path.join(ROOT, "FormS6Input.xlsx")
OUT = os.path.join(ROOT, "outputs", "web", "inputs.json")

VARS = ["CP", "Rmax", "VT", "WSP", "CF", "FFP", "Quantile"]
# 0-based column offsets for each category block (col A = index 0)
CAT_COLS = {"cat1": 1, "cat3": 9, "cat5": 17}


def main():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["SA all Variables"]
    rows = list(ws.iter_rows(min_row=3, values_only=True))  # data starts row 3

    out = {"vars": VARS, "cat1": [], "cat3": [], "cat5": []}
    for r in rows:
        if r[0] is None:
            continue
        vec = int(r[0])
        for cat, c0 in CAT_COLS.items():
            rec = {"vector": vec}
            for i, name in enumerate(VARS):
                v = r[c0 + i]
                rec[name] = round(float(v), 6) if v is not None else None
            out[cat].append(rec)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w") as f:
        json.dump(out, f)

    for cat in ("cat1", "cat3", "cat5"):
        n = len(out[cat])
        s = out[cat][0]
        print(f"{cat}: {n} vectors  (vector1: CP={s['CP']} Rmax={s['Rmax']} "
              f"VT={s['VT']} WSP={s['WSP']} CF={s['CF']} FFP={s['FFP']})")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
