#!/usr/bin/env python3
"""Powell (PDE) Uncertainty-Analysis precompute — Option 1 (faithful EPR).

Runs the Powell wind-field over the 6 one-variable-at-a-time UA worksheets
("UA for CP/Rmax/VT/WSP/CF/FFP") of FormS6Input.xlsx and records, for each
input vector, the scalar output metric used for SA/UA: the MEAN PEAK WIND over
the 682 land vertices. From these the viewer computes the faithful EPR per
variable as the variance share Var(Y when only X_i varies) / Var(Y_SA).

  6 sheets x 3 categories x 100 vectors = 1800 PDE solves (~70 min on MPS).

Output: outputs/web/powell_ua.json
  { "unit":"mph", "vars":[...],
    "cp":  {"cat1":[100], "cat3":[100], "cat5":[100]}, "rmax": {...}, ... }
"""
import os, sys, json, time
import torch
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import windfield_grid as W  # reuses make_args/peak_winds/constants + torch device

ROOT = os.path.dirname(HERE)
XLSX = os.path.join(ROOT, "FormS6Input.xlsx")
OUT = os.path.join(ROOT, "outputs", "web", "powell_ua.json")

# UA worksheet name -> short key
UA_SHEETS = {
    "cp": "UA for CP", "rmax": "UA for Rmax", "vt": "UA for VT",
    "wsp": "UA for WSP", "cf": "UA for CF", "ffp": "UA for FFP",
}
VARS = ["CP", "Rmax", "VT", "WSP", "CF", "FFP", "Quantile"]
CAT_COLS = {"cat1": 1, "cat3": 9, "cat5": 17}   # 0-based col offsets (col A = 0)


def read_sheet(wb, name):
    ws = wb[name]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    out = {"cat1": [], "cat3": [], "cat5": []}
    for r in rows:
        if r[0] is None:
            continue
        for cat, c0 in CAT_COLS.items():
            rec = {VARS[i]: (float(r[c0 + i]) if r[c0 + i] is not None else None)
                   for i in range(len(VARS))}
            out[cat].append(rec)
    return out


def main():
    device = W.H.device_select()
    grid = json.load(open(W.GRID))
    pts = grid["points"]
    land_idx = [i for i, p in enumerate(pts) if p["land"]]
    ew = torch.tensor([p["ew"] for p in pts], dtype=torch.float32, device=device)
    ns = torch.tensor([p["ns"] for p in pts], dtype=torch.float32, device=device)
    hours_t = torch.arange(0.0, W.T_MAX + W.T_DT / 2, W.T_DT, dtype=torch.float32, device=device)
    land_t = torch.tensor(land_idx, dtype=torch.long, device=device)

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    out = {"unit": "mph", "metric": "mean peak wind over 682 land vertices",
           "vars": ["CP", "Rmax", "VT", "WSP", "CF", "FFP"]}

    total = len(UA_SHEETS) * 3 * 100
    done, t0 = 0, time.time()
    for key, sheet in UA_SHEETS.items():
        data = read_sheet(wb, sheet)
        out[key] = {}
        for cat in ("cat1", "cat3", "cat5"):
            ys = []
            for rec in data[cat]:
                pw = W.peak_winds(rec, ew, ns, hours_t, device)   # (840,) mph
                ys.append(round(float(pw[land_t].mean()), 3))
                done += 1
                if done % 50 == 0 or done == total:
                    el = time.time() - t0
                    print(f"  {done}/{total} ({el:.0f}s, {el/done:.2f}s/solve, "
                          f"ETA {el/done*(total-done):.0f}s)", flush=True)
            out[key][cat] = ys
        print(f"{sheet}: done", flush=True)

    json.dump(out, open(OUT, "w"))
    print(f"Wrote {OUT} in {time.time()-t0:.0f}s", flush=True)


if __name__ == "__main__":
    main()
